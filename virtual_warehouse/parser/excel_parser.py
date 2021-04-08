"""Parser of Excel data files."""
from openpyxl import load_workbook
from xlrd import open_workbook

from virtual_warehouse.parser.data_model import (
    Inventory,
    Item,
    ItemUnit,
    Location,
    Order,
)
from virtual_warehouse.parser.utils import estimate_sheet_type


class Document:
    """Document class which loads xls or xlsx file and parse different data objects."""

    def __init__(self, file_path):
        # Determines backend for loading documents (xlsx files uses openpyxl)
        self.is_xlsx = Document.check_xlsx(file_path)
        self.locations = {}
        self.items = {}
        self.balance = {}
        self.orders = {}
        if self.is_xlsx:
            # self.doc = load_workbook(data_path, read_only=True, keep_links=False)
            self.doc = open_workbook(file_path)
        else:
            self.doc = open_workbook(file_path)

    @staticmethod
    def check_xlsx(file_path):
        """Check if document is .xlsx document (required for openpyxl library)."""
        return file_path[-5:] == ".xlsx"

    @staticmethod
    def get_sheet_names(file_path):
        """Get names of all sheets in document."""
        if Document.check_xlsx(file_path):
            doc = load_workbook(file_path, read_only=True, keep_links=False)
            names = doc.sheetnames
        else:
            doc = open_workbook(file_path, on_demand=True)
            names = doc.sheet_names()
        return [[n, estimate_sheet_type(n)] for n in names]

    def parse_locations(self, sheet_name="LOCATIONmaster"):
        """Parse LOCATIONmaster sheet."""
        sheet = self.doc.sheet_by_name(sheet_name)
        for row in range(1, sheet.nrows):
            location_id = str(sheet.cell(row, 0).value)
            if not location_id:
                continue

            self.locations[location_id] = Location.create(
                *(sheet.cell(row, i).value for i in range(11))
            )

        return self.locations

    def parse_coordinates(self, sheet_name="XYZ_coordinates"):
        """Parse XYZ_coordinates sheet."""
        sheet = self.doc.sheet_by_name(sheet_name)
        for row in range(1, sheet.nrows):
            location_id = str(sheet.cell(row, 0).value)
            if not location_id:
                continue

            self.locations[location_id].set_coord(
                *(sheet.cell(row, i).value for i in range(1, 4))
            )

        return self.locations

    def parse_items(self, sheet_name="ITEMmaster"):
        """Parse ITEMmaster sheet."""
        sheet = self.doc.sheet_by_name(sheet_name)
        for row in range(1, sheet.nrows):
            item_id, description, gtype, zone = (
                sheet.cell(row, i).value for i in range(4)
            )
            item_id = str(item_id)
            if not item_id:
                continue

            unit_levels = []
            for col in range(4, sheet.ncols, 8):
                unit_levels.append(
                    ItemUnit.create(
                        f"{item_id}-u{col}",
                        *(sheet.cell(row, col + i).value for i in range(8)),
                    )
                )
            self.items[item_id] = Item.create(
                item_id, description, gtype, zone, unit_levels[0], unit_levels
            )

        return self.items

    def parse_inventory_balance(self, sheet_name="Inventory Ballance"):
        """Parse Inventory Balance sheet  ('balance' in final version, most likely)."""
        sheet = self.doc.sheet_by_name(sheet_name)
        for row in range(1, sheet.nrows):
            date, location_id = sheet.cell(row, 0).value, sheet.cell(row, 1).value
            location_id = str(location_id)
            if not date:
                continue

            if date not in self.balance:
                self.balance[date] = {}
            self.balance[date][location_id] = Inventory.create(
                *(sheet.cell(row, i).value for i in range(10))
            )

        return self.balance

    def parse_orders(self, sheet_name="Order"):
        """Parse Order sheet."""
        sheet = self.doc.sheet_by_name(sheet_name)
        for row in range(1, sheet.nrows):
            order_id = str(sheet.cell(row, 0).value)
            if not order_id:
                continue

            if order_id in self.orders:
                self.orders[order_id].add_item(
                    *(sheet.cell(row, i).value for i in range(7, 11))
                )
            else:
                self.orders[order_id] = Order.create(
                    *(sheet.cell(row, i).value for i in range(11))
                )

        return self.orders

    def parse_document(self):
        """Parse the whole document."""
        locations = self.parse_locations()
        locations = self.parse_coordinates()
        items = self.parse_items()
        balance = self.parse_inventory_balance()
        orders = self.parse_orders()

        return locations, items, balance, orders
